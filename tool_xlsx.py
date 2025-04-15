from openpyxl import load_workbook



# import pandas as pd

# # 用 openpyxl 加载工作簿
# wb = load_workbook("原文件.xlsx")
# ws = wb.active

# # 用 pandas 读取数据（假设数据在A1:C3区域）
# df = pd.DataFrame(ws.values).iloc[0:3, 0:3]

# # 处理数据（示例：第一行第一列改为新值）
# df.iloc[0,0] = "新值"

# # 将数据写回原位置（不破坏样式）
# for row in df.itertuples(index=False):
#     for col_idx, value in enumerate(row, start=1):
#         ws.cell(row=row.Index+1, column=col_idx, value=value)

# wb.save("原文件.xlsx")
def 加载会话历史表(文件名):
    # return load_workbook(文件名)
    template = load_workbook(文件名, rich_text=True)
    sheet = template["会话表"]
    return sheet


def 写入新增健康档案记录(消息列表):
    '''
    涉及唯一性项识别合并问题：
    '''
    pass





if __name__ == "__main__":
    print(加载会话历史表("原文件.xlsx"))